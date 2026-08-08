"""Excel eksport/import yordamchilari."""

from __future__ import annotations

import io
from decimal import Decimal
from typing import Any, Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F6FEB")
HEADER_FONT = Font(color="FFFFFF", bold=True)

#: Katalog import/eksport ustunlari (o'zbekcha sarlavhalar bilan)
PRODUCT_COLUMNS: list[tuple[str, str]] = [
    ("sku", "SKU"),
    ("name", "Nomi"),
    ("category", "Kategoriya"),
    ("diameter_mm", "Diametr (mm)"),
    ("length_mm", "Uzunlik (mm)"),
    ("implant_type", "Turi"),
    ("connection_type", "Ulanish"),
    ("price_usd", "Narx (USD)"),
    ("min_stock", "Minimal qoldiq"),
    ("qty", "Qoldiq"),
]


def _style_header(ws, headers: Sequence[str]) -> None:
    ws.append(list(headers))
    for index, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws, headers: Sequence[str]) -> None:
    for index, header in enumerate(headers, start=1):
        longest = len(str(header))
        for row in range(2, min(ws.max_row + 1, 500)):
            value = ws.cell(row=row, column=index).value
            longest = max(longest, len(str(value)) if value is not None else 0)
        ws.column_dimensions[get_column_letter(index)].width = min(48, longest + 3)


def _to_cell(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if hasattr(value, "value") and not isinstance(value, (int, float, str)):
        return value.value
    return value


def build_workbook(
    sheets: dict[str, tuple[Sequence[tuple[str, str]], Iterable[dict]]]
) -> io.BytesIO:
    """Umumiy eksport: {varaq nomi: ([(kalit, sarlavha)], qatorlar)}."""
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, (columns, rows) in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        _style_header(ws, [title for _, title in columns])
        for row in rows:
            ws.append([_to_cell(row.get(key)) for key, _ in columns])
        _autosize(ws, [title for _, title in columns])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def build_products_workbook(products: Iterable[Any], categories: Iterable[Any]) -> io.BytesIO:
    rows = []
    for product in products:
        rows.append(
            {
                "sku": product.sku,
                "name": product.name,
                "category": product.category_name,
                "diameter_mm": product.diameter_mm,
                "length_mm": product.length_mm,
                "implant_type": product.implant_type,
                "connection_type": product.connection_type,
                "price_usd": product.price_usd,
                "min_stock": product.min_stock,
                "qty": product.qty,
            }
        )

    stream = build_workbook({"Katalog": (PRODUCT_COLUMNS, rows)})

    # Kategoriyalar ro'yxatini ham qo'shamiz — import qilishda nom aniq bo'lsin
    wb = load_workbook(stream)
    ws = wb.create_sheet("Kategoriyalar")
    _style_header(ws, ["Kod", "Nomi"])
    for category in categories:
        ws.append([category.code, category.name_uz])
    _autosize(ws, ["Kod", "Nomi"])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def parse_products_workbook(stream: io.BytesIO) -> list[dict]:
    """Excel'dan mahsulotlarni o'qiydi. Sarlavhalar `PRODUCT_COLUMNS` bo'yicha."""
    wb = load_workbook(stream, data_only=True)
    ws = wb["Katalog"] if "Katalog" in wb.sheetnames else wb.worksheets[0]

    header_row = [str(c.value or "").strip() for c in ws[1]]
    title_to_key = {title: key for key, title in PRODUCT_COLUMNS}
    mapping: dict[int, str] = {}
    for index, title in enumerate(header_row):
        key = title_to_key.get(title) or (title if title in dict(PRODUCT_COLUMNS) else None)
        if key:
            mapping[index] = key

    if "sku" not in mapping.values() or "name" not in mapping.values():
        raise ValueError(
            "Faylda «SKU» va «Nomi» ustunlari bo'lishi shart. "
            "Avval katalogni Excel'ga yuklab olib, o'sha shaklda to'ldiring."
        )

    result: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record: dict = {}
        for index, key in mapping.items():
            record[key] = row[index] if index < len(row) else None
        if not record.get("sku") or not record.get("name"):
            continue
        record["sku"] = str(record["sku"]).strip()
        record["name"] = str(record["name"]).strip()
        for numeric in ("diameter_mm", "length_mm", "price_usd"):
            if record.get(numeric) in ("", None):
                record[numeric] = None
            else:
                record[numeric] = Decimal(str(record[numeric]).replace(",", "."))
        record["min_stock"] = int(record.get("min_stock") or 0)
        for text in ("implant_type", "connection_type", "category"):
            value = record.get(text)
            record[text] = str(value).strip() if value not in (None, "") else None
        result.append(record)

    if not result:
        raise ValueError("Faylda birorta ham mahsulot topilmadi")
    return result
